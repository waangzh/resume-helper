#!/usr/bin/env python
"""Export job recommendations from CSV or JSON to a formatted Excel workbook."""

from __future__ import annotations

import argparse
import csv
import json
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADERS = [
    "岗位名称",
    "公司名称",
    "城市",
    "平台",
    "学历要求",
    "技术关键词",
    "岗位详情页",
    "来源",
    "检索日期",
    "推荐理由",
    "投递要求",
    "状态",
    "备注",
]

KEY_MAP = {
    "岗位名称": ["岗位名称", "title", "job_title", "name"],
    "公司名称": ["公司名称", "company", "company_name"],
    "城市": ["城市", "city", "location"],
    "平台": ["平台", "platform"],
    "学历要求": ["学历要求", "education_requirement", "education"],
    "技术关键词": ["技术关键词", "keywords", "skills"],
    "岗位详情页": ["岗位详情页", "url", "link", "job_url"],
    "来源": ["来源", "source"],
    "检索日期": ["检索日期", "retrieved_at", "date"],
    "推荐理由": ["推荐理由", "reason"],
    "投递要求": ["投递要求", "apply_requirements"],
    "状态": ["状态", "status"],
    "备注": ["备注", "note", "notes"],
}


def load_records(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".json":
        data = json.loads(path.read_text(encoding="utf-8-sig"))
        if isinstance(data, dict):
            data = data.get("jobs", data.get("recommendations", []))
        if not isinstance(data, list):
            raise ValueError("JSON input must be a list or contain jobs/recommendations list.")
        return [dict(item) for item in data]
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def pick(record: dict[str, Any], header: str) -> str:
    for key in KEY_MAP[header]:
        if key in record and record[key] not in (None, ""):
            value = record[key]
            if isinstance(value, list):
                return "、".join(str(item) for item in value)
            return str(value)
    if header == "检索日期":
        return date.today().isoformat()
    if header == "状态":
        return "待用户确认"
    return ""


def export_excel(records: list[dict[str, Any]], output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "岗位推荐清单"
    ws.append(HEADERS)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    link_col = HEADERS.index("岗位详情页") + 1
    for record in records:
        row = [pick(record, header) for header in HEADERS]
        ws.append(row)
        link_cell = ws.cell(row=ws.max_row, column=link_col)
        if link_cell.value:
            link_cell.hyperlink = str(link_cell.value)
            link_cell.style = "Hyperlink"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        "A": 24,
        "B": 22,
        "C": 12,
        "D": 14,
        "E": 12,
        "F": 28,
        "G": 48,
        "H": 18,
        "I": 14,
        "J": 42,
        "K": 24,
        "L": 14,
        "M": 28,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> int:
    parser = argparse.ArgumentParser(description="Export job recommendations to a formatted .xlsx file.")
    parser.add_argument("--input", required=True, help="Input CSV or JSON recommendations file.")
    parser.add_argument("--output", required=True, help="Output .xlsx path.")
    args = parser.parse_args()

    export_excel(load_records(Path(args.input)), Path(args.output))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
